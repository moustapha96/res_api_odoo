# -*- coding: utf-8 -*-
from .main import *
import sys
import pdb
import logging
_logger = logging.getLogger(__name__)


import os
import xlsxwriter
import openpyxl

from odoo.http import request
from datetime import datetime, timedelta



class ExcelUpdateController(http.Controller):


    # @http.route('/api/update_excel', methods=['POST'],auth="none" ,type='http', cors="*", csrf=False)
    # def update_excel(self, **kw):
    #     current_dir = os.path.dirname(os.path.abspath(__file__))
    #     parent_dir = os.path.dirname(current_dir)

    #     # Chemin complet vers le fichier Excel dans le répertoire parent
    #     file_path = os.path.join(parent_dir, 'data.xlsx')
    #     data = json.loads(request.httprequest.data)
    #     # Données à ajouter (reçues via la requête POST)
    #     new_data = data.get('data')

    #     if not new_data:
    #         _logger.info("No data received")
    #         return request.make_response(
    #             json.dumps({"status": "error", "message": "Aucune donnée reçue"}),
    #             status=400,
    #             headers={'Content-Type': 'application/json'}
    #         )

    #     # Vérifier si le fichier existe
    #     if not os.path.exists(file_path):
    #         # Créer un nouveau fichier Excel avec xlsxwriter
    #         workbook = xlsxwriter.Workbook(file_path)
    #         worksheet = workbook.add_worksheet()

    #         # En-têtes de colonnes
    #         headers = ['Product Name', 'Date', 'User', 'Quantity', 'Phone', 'Email']
    #         for col_num, header in enumerate(headers):
    #             worksheet.write(0, col_num, header)

    #         # Ajouter les nouvelles données
    #         for row_num, row_data in enumerate(new_data, start=1):
    #             worksheet.write(row_num, 0, row_data.get('productName'))
    #             worksheet.write(row_num, 1, row_data.get('date'))
    #             worksheet.write(row_num, 2, row_data.get('user'))
    #             worksheet.write(row_num, 3, row_data.get('quantity', ''))  # Gérer l'absence de 'quantity'
    #             worksheet.write(row_num, 4, row_data.get('phone', ''))     # Gérer l'absence de 'phone'
    #             worksheet.write(row_num, 5, row_data.get('email', ''))     # Gérer l'absence de 'email'

    #         workbook.close()
           
    #         return request.make_response(
    #             json.dumps({"status": "success", "message": "Fichier Excel créé et mis à jour avec succès"}),
    #             status=200,
    #             headers={'Content-Type': 'application/json'}
    #         )
    #     else:
    #         # Mettre à jour le fichier existant avec openpyxl
    #         workbook = openpyxl.load_workbook(file_path)
    #         worksheet = workbook.active

    #         # Trouver la première ligne vide
    #         row_num = worksheet.max_row + 1

    #         # Ajouter les nouvelles données
    #         for row_data in new_data:
    #             worksheet.cell(row=row_num, column=1, value=row_data.get('productName'))
    #             worksheet.cell(row=row_num, column=2, value=row_data.get('date'))
    #             worksheet.cell(row=row_num, column=3, value=row_data.get('user'))
    #             worksheet.cell(row=row_num, column=4, value=row_data.get('quantity', ''))
    #             worksheet.cell(row=row_num, column=5, value=row_data.get('phone', ''))
    #             worksheet.cell(row=row_num, column=6, value=row_data.get('email', ''))
    #             row_num += 1

    #         workbook.save(file_path)
          
    #         return request.make_response(
    #             json.dumps({"status": "success", "message": "Fichier Excel mis à jour avec succès"}),
    #             status=200,
    #             headers={'Content-Type': 'application/json'}
    #         )


    @http.route('/api/create_leads', methods=['POST'], auth="none", type='http', cors="*", csrf=False)
    def create_leads(self, **kw):
        data = json.loads(request.httprequest.data)
        leads_data = data.get('data')

        if not leads_data or not isinstance(leads_data, list):
            _logger.info("No lead data received or data is not a list")
            return request.make_response(
                json.dumps({"status": "error", "message": "Aucune donnée de lead reçue ou format invalide"}),
                status=400,
                headers={'Content-Type': 'application/json'}
            )

        if not request.env.user or request.env.user._is_public():
            admin_user = request.env.ref('base.user_admin')
            request.env = request.env(user=admin_user.id)


        user_ip = request.httprequest.remote_addr
        _logger.info(f"User IP: {user_ip}")
        

        Lead = request.env['crm.lead'].sudo()
        Partner = request.env['res.partner'].sudo()  # Modèle des partenaires
        created_leads = []
        
        # Récupérer les emails existants pour éviter les doublons
        # existing_emails = Lead.search([('email_from', 'in', [lead.get('email') for lead in leads_data])])
       
        tag_produit = request.env['crm.tag'].sudo().search([('name', '=', 'Produit')], limit=1)
    

        for lead in leads_data:
            # Vérification des champs obligatoires
            if 'productName' not in lead or 'email' not in lead or 'type' not in lead:
                _logger.info("Missing required fields in lead data")
                continue  # Ignorer ce lead si les champs obligatoires manquent

            existing_lead = Lead.search([
                ('email_from', '=', lead['email']),
                ('name', '=', lead['productName']),
                ('date_deadline', '>=', datetime.now().date()),  # Vérifie que la date est aujourd'hui ou plus tard
                ('date_deadline', '<=', datetime.now().date())   # Vérifie que la date est aujourd'hui
            ], limit=1)

            # Vérifier si un lead avec le même email existe déjà
            if existing_lead:
                _logger.info(f"Lead with product '{lead['productName']}', email '{lead['email']}' and today's date already exists. Skipping.")
                continue 

            # Vérifier si un partenaire existe avec le même email
            partner = Partner.search([('email', '=', lead['email'])], limit=1)
            partner_id = partner.id if partner else False  # Lier le partenaire si trouvé

            # Définir la date limite en fonction du type de lead
            if lead['type'] == 'order':
                date_deadline = datetime.now() + timedelta(days=7)  # 7 jours pour les commandes
            elif lead['type'] == 'preorder':
                date_deadline = datetime.now() + timedelta(days=60)  # 60 jours pour les précommandes
            else:
                date_deadline = None  # Pas de date limite pour les autres types

            # Création du lead
            new_lead = Lead.create({
                'name': lead['productName'],  # Nom du lead (utilisation de productName)
                'email_from': lead['email'],  # Email du lead
                'phone': lead.get('phone'),  # Téléphone (optionnel)
                'user_id': request.env.user.id,  # Assigner à l'utilisateur courant
                'description': f"Date: {lead['date']}, User: {lead['user']}, Type: {lead['type']}",  # Description optionnelle
                'date_deadline': date_deadline,  # Date limite en fonction du type
                'partner_id': partner_id,
                'expected_revenue': lead['price'],
                'tag_ids': [(6, 0, [tag_produit.id])] if tag_produit else [],
                'location': lead.get('adresse'),
            })
            created_leads.append(new_lead.id)

        return request.make_response(
            json.dumps({"status": "success", "message": "Leads créés avec succès", "lead_ids": created_leads}),
            status=201,
            headers={'Content-Type': 'application/json'}
        )